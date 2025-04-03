#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel to JSON converter with merged cell and metadata detection.
Handles hierarchical data and detects metadata sections at the top of Excel files.
"""

import pandas as pd
import json
import openpyxl
from pathlib import Path
import os
import logging
import hashlib
import pickle
from typing import Dict, List, Tuple, Optional, Any, Union

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.FileHandler("excel_processing.log"), logging.StreamHandler()]
)
logger = logging.getLogger("excel_processor")


class ExcelProcessingError(Exception):
    """Base exception for Excel processing errors"""
    pass


class MergeMapError(ExcelProcessingError):
    """Error building the merge map"""
    pass


class MetadataExtractionError(ExcelProcessingError):
    """Error extracting metadata"""
    pass


class DataExtractionError(ExcelProcessingError):
    """Error extracting data"""
    pass


def load_workbook(excel_file: str) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    """
    Load an Excel workbook and return the workbook and active sheet
    """
    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        return wb, sheet
    except Exception as e:
        logger.error(f"Failed to load workbook {excel_file}: {str(e)}")
        raise ExcelProcessingError(f"Failed to load workbook: {str(e)}") from e


def build_merge_map(sheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict:
    """
    Create a mapping of merged regions in the sheet
    """
    try:
        merge_map = {}
        for merged_range in sheet.merged_cells.ranges:
            # Find the value from the top-left cell
            top_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
            
            # Record this merge in our map
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merge_map[(row, col)] = {
                        'value': top_value,
                        'origin': (merged_range.min_row, merged_range.min_col),
                        'range': str(merged_range)
                    }
        
        logger.info(f"Found {len(sheet.merged_cells.ranges)} merged regions")
        return merge_map
    except Exception as e:
        logger.error(f"Failed to build merge map: {str(e)}")
        raise MergeMapError(f"Failed to build merge map: {str(e)}") from e


def extract_metadata(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                    merge_map: Dict, 
                    max_metadata_rows: int = 6) -> Tuple[Dict, int]:
    """
    Extract metadata section from the top of the Excel file
    """
    try:
        metadata = {}
        metadata_rows = 0
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Check for potential metadata header sections (large merged cells at the top)
        for merged_range in sheet.merged_cells.ranges:
            # If there's a merged region at the top spanning multiple columns
            if (merged_range.min_row <= 3 and  # In first few rows
                (merged_range.max_col - merged_range.min_col + 1) > 2):  # Spans several columns
                
                metadata_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
                if metadata_value:
                    key = f"header_r{merged_range.min_row}"
                    metadata[key] = metadata_value
                    metadata_rows = max(metadata_rows, merged_range.max_row)
        
        # Look for metadata in the first few rows (labels, dates, document info)
        for row in range(1, min(max_metadata_rows + 1, max_row + 1)):
            row_has_data = False
            row_data = {}
            
            for col in range(1, max_col + 1):
                value = None
                # Get value accounting for merged cells
                if (row, col) in merge_map:
                    # If this cell is part of a large merged region already processed as a header, skip
                    if any(merge_map[(row, col)]['origin'] == (r.min_row, r.min_col) 
                          for r in sheet.merged_cells.ranges 
                          if r.min_row <= 3 and (r.max_col - r.min_col + 1) > 2):
                        continue
                    value = merge_map[(row, col)]['value']
                else:
                    value = get_typed_cell_value(sheet.cell(row, col))
                
                if value is not None:
                    # Try to get column header from first row if this is not the first row
                    col_header = None
                    if row > 1:
                        col_header = sheet.cell(1, col).value
                    
                    # Use column header or column letter
                    key = col_header if col_header else openpyxl.utils.get_column_letter(col)
                    row_data[key] = value
                    row_has_data = True
            
            if row_has_data:
                metadata[f"row_{row}"] = row_data
                metadata_rows = max(metadata_rows, row)
        
        logger.info(f"Detected metadata section up to row {metadata_rows}")
        return metadata, metadata_rows
    except Exception as e:
        logger.error(f"Failed to extract metadata: {str(e)}")
        raise MetadataExtractionError(f"Failed to extract metadata: {str(e)}") from e


def identify_data_start(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                       metadata_rows: int, 
                       merge_map: Dict,
                       header_threshold: int = 3) -> int:
    """
    Determine where the main data begins after metadata
    """
    data_start_row = metadata_rows + 1
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Look for a clear header row (typically follows metadata and has values across columns)
    for row in range(data_start_row, min(data_start_row + 5, max_row + 1)):
        values_in_row = 0
        for col in range(1, max_col + 1):
            if (row, col) in merge_map:
                if merge_map[(row, col)]['value'] is not None:
                    values_in_row += 1
            elif sheet.cell(row, col).value is not None:
                values_in_row += 1
        
        # If we find a row with several populated cells, it's likely a header
        threshold = max(header_threshold, max_col / 3)
        if values_in_row > threshold:  # Either header_threshold or 1/3 of columns have values
            data_start_row = row
            break
    
    logger.info(f"Main data starts at row {data_start_row}")
    return data_start_row


def get_typed_cell_value(cell: openpyxl.cell.cell.Cell) -> Any:
    """
    Extract cell value with appropriate type information
    """
    if cell.value is None:
        return None
        
    if cell.data_type == 'n':  # Number
        return float(cell.value) if isinstance(cell.value, float) else int(cell.value)
    elif cell.data_type == 'd':  # Date
        return cell.value.isoformat() if hasattr(cell.value, 'isoformat') else str(cell.value)
    elif cell.data_type == 'b':  # Boolean
        return bool(cell.value)
    
    # Default to string for other types
    return str(cell.value) if cell.value is not None else None


def get_cell_value(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                  row: int, col: int, 
                  merge_map: Dict) -> Any:
    """
    Helper function to get cell value, considering merged cells
    """
    excel_row = row + 1  # 1-based indexing in openpyxl
    excel_col = col + 1
    if (excel_row, excel_col) in merge_map:
        return merge_map[(excel_row, excel_col)]['value']
    else:
        return get_typed_cell_value(sheet.cell(excel_row, excel_col))


def extract_hierarchical_data(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                             merge_map: Dict, 
                             data_start_row: int,
                             df_headers: List[str],
                             chunk_size: int = 1000,
                             include_empty: bool = False) -> List[Dict]:
    """
    Extract hierarchical data from the Excel sheet
    """
    try:
        max_row = sheet.max_row
        max_col = sheet.max_column
        hierarchical_data = []
        rows_processed = set()
        
        # Process in chunks to handle large files
        for chunk_start in range(data_start_row-1, max_row, chunk_size):
            chunk_end = min(chunk_start + chunk_size, max_row)
            logger.debug(f"Processing rows {chunk_start+1} to {chunk_end}")
            
            # Process row by row
            row_idx = chunk_start
            while row_idx < chunk_end:
                if row_idx in rows_processed:
                    row_idx += 1
                    continue
                    
                row_data = {}
                skip_rows = 1
                
                # Process each column
                for col_idx in range(max_col):
                    value = get_cell_value(sheet, row_idx, col_idx, merge_map)
                    
                    # Skip empty values if configured
                    if value is None and not include_empty:
                        continue
                    
                    # Check if this is a merged cell origin
                    cell_key = (row_idx + 1, col_idx + 1)
                    if cell_key in merge_map and merge_map[cell_key]['origin'] == cell_key:
                        # This is a merged cell origin - handle specially
                        # Find how many rows this spans
                        for m_range in sheet.merged_cells.ranges:
                            if m_range.min_row == row_idx + 1 and m_range.min_col == col_idx + 1:
                                span = m_range.max_row - m_range.min_row + 1
                                if span > skip_rows:
                                    skip_rows = span
                        
                        # For columns with multiple data points within a merged parent
                        if col_idx > 0:  # Not the first column
                            sub_values = []
                            for sub_row in range(row_idx, row_idx + skip_rows):
                                if col_idx + 1 < max_col:  # Check next column exists
                                    sub_val = get_cell_value(sheet, sub_row, col_idx + 1, merge_map)
                                    if sub_val is not None or include_empty:
                                        sub_values.append(sub_val)
                            
                            # Add to row data with subpoints
                            col_name = df_headers[col_idx] if col_idx < len(df_headers) else f"Column_{col_idx}"
                            row_data[col_name] = {
                                'value': value,
                                'sub_values': sub_values
                            }
                        else:
                            # Add as regular value
                            col_name = df_headers[col_idx] if col_idx < len(df_headers) else f"Column_{col_idx}"
                            row_data[col_name] = value
                    else:
                        # Regular cell
                        col_name = df_headers[col_idx] if col_idx < len(df_headers) else f"Column_{col_idx}"
                        row_data[col_name] = value
                
                # Add this row to our results if it has any non-None values
                if any(v is not None for v in row_data.values()):
                    hierarchical_data.append(row_data)
                
                # Mark rows as processed
                for i in range(skip_rows):
                    rows_processed.add(row_idx + i)
                
                row_idx += skip_rows
        
        logger.info(f"Processed {len(hierarchical_data)} hierarchical records")
        return hierarchical_data
    except Exception as e:
        logger.error(f"Failed to extract hierarchical data: {str(e)}")
        raise DataExtractionError(f"Failed to extract hierarchical data: {str(e)}") from e


def create_result_structure(metadata: Dict, hierarchical_data: List[Dict]) -> Dict:
    """
    Create the final result structure with metadata and data
    """
    return {
        "metadata": metadata,
        "data": hierarchical_data
    }


def write_json_output(result: Dict, json_file: str) -> None:
    """
    Write the result structure to a JSON file
    """
    try:
        with open(json_file, 'w') as f:
            json.dump(result, f, indent=2)
        logger.info(f"Data written to {json_file}")
    except Exception as e:
        logger.error(f"Failed to write JSON output: {str(e)}")
        raise ExcelProcessingError(f"Failed to write JSON output: {str(e)}") from e


def convert_hierarchical_excel(excel_file: str, json_file: str, config: Dict = None) -> Dict:
    """
    Convert Excel with complex merged cells to properly structured JSON
    Specifically handles hierarchical data where parent cells are merged across multiple rows
    Also detects metadata sections at the beginning of the Excel file
    
    Args:
        excel_file: Path to Excel file
        json_file: Path for JSON output
        config: Dictionary with optional parameters:
            - metadata_max_rows: Max rows to check for metadata (default: 6)
            - header_detection_threshold: Min values to consider a header row (default: 3)
            - sheet_name: Specific sheet to process (default: active sheet)
            - include_empty_cells: Whether to include null values (default: False)
            - chunk_size: Number of rows to process at once (default: 1000)
    """
    logger.info(f"Processing hierarchical data from {excel_file}")
    
    # Parse configuration
    config = config or {}
    metadata_max_rows = config.get('metadata_max_rows', 6)
    header_threshold = config.get('header_detection_threshold', 3)
    sheet_name = config.get('sheet_name')
    include_empty = config.get('include_empty_cells', False)
    chunk_size = config.get('chunk_size', 1000)
    
    # Load workbook
    wb, sheet = load_workbook(excel_file)
    
    # Use specified sheet if provided
    if sheet_name and sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        logger.info(f"Using sheet: {sheet_name}")
    
    # Extract structure and metadata
    merge_map = build_merge_map(sheet)
    metadata, metadata_rows = extract_metadata(sheet, merge_map, metadata_max_rows)
    data_start_row = identify_data_start(sheet, metadata_rows, merge_map, header_threshold)
    
    # Read column names with pandas - more reliable for complex headers
    df = pd.read_excel(excel_file, header=data_start_row-1, sheet_name=sheet.title if sheet_name else 0)
    df_headers = list(df.columns)
    
    # Process hierarchical data
    hierarchical_data = extract_hierarchical_data(
        sheet, merge_map, data_start_row, df_headers, chunk_size, include_empty
    )
    
    # Combine and output
    result = create_result_structure(metadata, hierarchical_data)
    
    if json_file:
        write_json_output(result, json_file)
    
    logger.info(f"Processed {len(hierarchical_data)} hierarchical records with metadata")
    return result


def get_file_hash(filepath: str) -> str:
    """
    Calculate MD5 hash of a file to detect changes
    """
    hasher = hashlib.md5()
    with open(filepath, 'rb') as f:
        buf = f.read(65536)
        while len(buf) > 0:
            hasher.update(buf)
            buf = f.read(65536)
    return hasher.hexdigest()


def batch_process_excel_files(input_dir: str, output_dir: str, config: Dict = None, 
                             use_cache: bool = True, cache_dir: str = '.cache') -> Dict:
    """
    Process all Excel files in a directory and convert them to JSON with metadata detection
    
    Args:
        input_dir: Directory containing Excel files
        output_dir: Directory to write JSON output
        config: Configuration for Excel processing (see convert_hierarchical_excel)
        use_cache: Whether to use caching to avoid reprocessing unchanged files
        cache_dir: Directory to store cache files
    """
    input_path = Path(input_dir)
    output_path = Path(output_dir)
    os.makedirs(output_path, exist_ok=True)
    
    if use_cache:
        os.makedirs(cache_dir, exist_ok=True)
    
    excel_files = list(input_path.glob("*.xlsx")) + list(input_path.glob("*.xls"))
    logger.info(f"Found {len(excel_files)} Excel files to process")
    
    results = {}
    for excel_file in excel_files:
        json_file = output_path / f"{excel_file.stem}.json"
        
        try:
            process_file = True
            result = None
            
            # Check cache if enabled
            if use_cache:
                file_hash = get_file_hash(str(excel_file))
                cache_file = os.path.join(cache_dir, f"{excel_file.stem}_{file_hash}.pkl")
                
                if os.path.exists(cache_file):
                    logger.info(f"Using cached version for {excel_file.name}")
                    with open(cache_file, 'rb') as f:
                        result = pickle.load(f)
                    process_file = False
            
            # Process the file if needed
            if process_file:
                result = convert_hierarchical_excel(str(excel_file), str(json_file), config)
                
                # Cache the result if caching is enabled
                if use_cache:
                    file_hash = get_file_hash(str(excel_file))
                    cache_file = os.path.join(cache_dir, f"{excel_file.stem}_{file_hash}.pkl")
                    with open(cache_file, 'wb') as f:
                        pickle.dump(result, f)
            elif json_file:
                # Write cached result to JSON
                with open(json_file, 'w') as f:
                    json.dump(result, f, indent=2)
            
            results[excel_file.name] = {
                "status": "success",
                "output_file": str(json_file),
                "metadata_rows": len(result["metadata"]),
                "data_rows": len(result["data"])
            }
        except Exception as e:
            logger.error(f"Error processing {excel_file}: {str(e)}")
            results[excel_file.name] = {
                "status": "error",
                "error": str(e)
            }
    
    # Write summary report
    summary_file = output_path / "processing_summary.json"
    with open(summary_file, 'w') as f:
        json.dump(results, f, indent=2)
    
    logger.info(f"Batch processing complete. Summary written to {summary_file}")
    return results


def process_workbook(excel_file: str, json_file: str, sheet_names: List[str] = None, config: Dict = None) -> Dict:
    """
    Process multiple sheets from a workbook
    
    Args:
        excel_file: Path to Excel file
        json_file: Path for JSON output
        sheet_names: List of sheet names to process (None for all sheets)
        config: Configuration for Excel processing
    """
    wb = openpyxl.load_workbook(excel_file)
    
    # Process specified sheets or all sheets
    sheets_to_process = sheet_names or wb.sheetnames
    logger.info(f"Processing sheets: {', '.join(sheets_to_process)}")
    
    result = {"sheets": {}}
    for sheet_name in sheets_to_process:
        if sheet_name in wb.sheetnames:
            logger.info(f"Processing sheet: {sheet_name}")
            sheet_config = config.copy() if config else {}
            sheet_config['sheet_name'] = sheet_name
            
            # Don't write individual JSON files for sheets
            sheet_data = convert_hierarchical_excel(excel_file, None, sheet_config)
            result["sheets"][sheet_name] = sheet_data
        else:
            logger.warning(f"Sheet not found: {sheet_name}")
    
    # Write combined result to JSON
    if json_file:
        with open(json_file, 'w') as f:
            json.dump(result, f, indent=2)
        logger.info(f"Multi-sheet data written to {json_file}")
    
    return result


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Convert Excel files with complex hierarchical data to JSON')
    subparsers = parser.add_subparsers(dest='command', help='Command to run')
    
    # Single file processing
    single_parser = subparsers.add_parser('single', help='Process a single Excel file')
    single_parser.add_argument('-i', '--input', required=True, help='Input Excel file')
    single_parser.add_argument('-o', '--output', required=True, help='Output JSON file')
    single_parser.add_argument('-s', '--sheet', help='Specific sheet to process')
    single_parser.add_argument('-m', '--metadata-rows', type=int, default=6, 
                              help='Maximum rows to check for metadata')
    single_parser.add_argument('-e', '--include-empty', action='store_true', 
                              help='Include empty cells in output')
    
    # Multi-sheet processing
    multi_parser = subparsers.add_parser('multi', help='Process multiple sheets in an Excel file')
    multi_parser.add_argument('-i', '--input', required=True, help='Input Excel file')
    multi_parser.add_argument('-o', '--output', required=True, help='Output JSON file')
    multi_parser.add_argument('-s', '--sheets', nargs='+', help='Sheets to process (default: all)')
    
    # Batch processing
    batch_parser = subparsers.add_parser('batch', help='Process all Excel files in a directory')
    batch_parser.add_argument('-i', '--input-dir', required=True, help='Input directory')
    batch_parser.add_argument('-o', '--output-dir', required=True, help='Output directory')
    batch_parser.add_argument('-c', '--cache', action='store_true', help='Use caching for unchanged files')
    batch_parser.add_argument('--cache-dir', default='.cache', help='Cache directory')
    
    args = parser.parse_args()
    
    if args.command == 'single':
        config = {
            'metadata_max_rows': args.metadata_rows,
            'include_empty_cells': args.include_empty
        }
        if args.sheet:
            config['sheet_name'] = args.sheet
        
        convert_hierarchical_excel(args.input, args.output, config)
    
    elif args.command == 'multi':
        process_workbook(args.input, args.output, args.sheets)
    
    elif args.command == 'batch':
        batch_process_excel_files(args.input_dir, args.output_dir, 
                                 use_cache=args.cache, cache_dir=args.cache_dir)
    
    else:
        print("Excel metadata processor module. Use --help for usage information.")
