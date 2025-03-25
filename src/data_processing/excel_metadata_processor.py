#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel to JSON converter with merged cell and metadata detection.
Handles hierarchical data and detects metadata sections at the top of Excel files.
"""

import pandas as pd
import json
import openpyxl
from pathlib import Path
import os


def convert_hierarchical_excel(excel_file, json_file):
    """
    Convert Excel with complex merged cells to properly structured JSON
    Specifically handles hierarchical data where parent cells are merged across multiple rows
    Also detects metadata sections at the beginning of the Excel file
    """
    print(f"Processing hierarchical data from {excel_file}")
    
    # First extract merge information
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    # Get dimensions
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Create a mapping of merged regions
    merge_map = {}
    for merged_range in sheet.merged_cells.ranges:
        # Find the value from the top-left cell
        top_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
        
        # Record this merge in our map
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merge_map[(row, col)] = {
                    'value': top_value,
                    'origin': (merged_range.min_row, merged_range.min_col),
                    'range': str(merged_range)
                }
    
    print(f"Found {len(sheet.merged_cells.ranges)} merged regions")
    
    # Detect metadata section at the top
    metadata = {}
    metadata_rows = 0
    
    # Check for potential metadata header sections (large merged cells at the top)
    for merged_range in sheet.merged_cells.ranges:
        # If there's a merged region at the top spanning multiple columns
        if (merged_range.min_row <= 3 and  # In first few rows
            (merged_range.max_col - merged_range.min_col + 1) > 2):  # Spans several columns
            
            metadata_value = sheet.cell(merged_range.min_row, merged_range.min_col).value
            if metadata_value:
                key = f"header_r{merged_range.min_row}"
                metadata[key] = metadata_value
                metadata_rows = max(metadata_rows, merged_range.max_row)
    
    # Look for metadata in the first few rows (labels, dates, document info)
    for row in range(1, min(6, max_row + 1)):  # Check first 5 rows
        row_has_data = False
        row_data = {}
        
        for col in range(1, max_col + 1):
            value = None
            # Get value accounting for merged cells
            if (row, col) in merge_map:
                # If this cell is part of a large merged region already processed as a header, skip
                if any(merge_map[(row, col)]['origin'] == (r.min_row, r.min_col) 
                      for r in sheet.merged_cells.ranges 
                      if r.min_row <= 3 and (r.max_col - r.min_col + 1) > 2):
                    continue
                value = merge_map[(row, col)]['value']
            else:
                value = sheet.cell(row, col).value
            
            if value is not None:
                # Try to get column header from first row if this is not the first row
                col_header = None
                if row > 1:
                    col_header = sheet.cell(1, col).value
                
                # Use column header or column letter
                key = col_header if col_header else openpyxl.utils.get_column_letter(col)
                row_data[key] = value
                row_has_data = True
        
        if row_has_data:
            metadata[f"row_{row}"] = row_data
            metadata_rows = max(metadata_rows, row)
    
    # Determine where to start the main data processing (after metadata)
    data_start_row = metadata_rows + 1
    # Look for a clear header row (typically follows metadata and has values across columns)
    for row in range(data_start_row, min(data_start_row + 5, max_row + 1)):
        values_in_row = 0
        for col in range(1, max_col + 1):
            if (row, col) in merge_map:
                if merge_map[(row, col)]['value'] is not None:
                    values_in_row += 1
            elif sheet.cell(row, col).value is not None:
                values_in_row += 1
        
        # If we find a row with several populated cells, it's likely a header
        if values_in_row > max(3, max_col / 3):  # Either 3 or 1/3 of columns have values
            data_start_row = row
            break
    
    print(f"Detected metadata section up to row {metadata_rows}")
    print(f"Main data starts at row {data_start_row}")
    
    # Read data with pandas - we'll still use this for column names
    df = pd.read_excel(excel_file, header=data_start_row-1)  # Assuming header is just before data
    
    # Helper function to get value, considering merged cells
    def get_cell_value(row, col):
        excel_row = row + 1  # 1-based indexing in openpyxl
        excel_col = col + 1
        if (excel_row, excel_col) in merge_map:
            return merge_map[(excel_row, excel_col)]['value']
        else:
            return sheet.cell(excel_row, excel_col).value
    
    # Create a properly structured hierarchical representation
    hierarchical_data = []
    rows_processed = set()
    
    # Process row by row to build hierarchical structure, starting after metadata/header
    for row_idx in range(data_start_row-1, max_row):  # Convert to 0-based for looping
        if row_idx in rows_processed:
            continue
            
        row_data = {}
        skip_rows = 1
        
        # Process each column
        for col_idx in range(max_col):
            value = get_cell_value(row_idx, col_idx)
            
            # Check if this is a merged cell origin
            cell_key = (row_idx + 1, col_idx + 1)
            if cell_key in merge_map and merge_map[cell_key]['origin'] == cell_key:
                # This is a merged cell origin - handle specially
                # Find how many rows this spans
                for m_range in sheet.merged_cells.ranges:
                    if m_range.min_row == row_idx + 1 and m_range.min_col == col_idx + 1:
                        span = m_range.max_row - m_range.min_row + 1
                        if span > skip_rows:
                            skip_rows = span
                
                # For columns with multiple data points within a merged parent
                if col_idx > 0:  # Not the first column
                    sub_values = []
                    for sub_row in range(row_idx, row_idx + skip_rows):
                        if col_idx + 1 < max_col:  # Check next column exists
                            sub_val = get_cell_value(sub_row, col_idx + 1)
                            if sub_val is not None:
                                sub_values.append(sub_val)
                    
                    # Add to row data with subpoints
                    col_name = df.columns[col_idx] if col_idx < len(df.columns) else f"Column_{col_idx}"
                    row_data[col_name] = {
                        'value': value,
                        'sub_values': sub_values
                    }
                else:
                    # Add as regular value
                    col_name = df.columns[col_idx] if col_idx < len(df.columns) else f"Column_{col_idx}"
                    row_data[col_name] = value
            else:
                # Regular cell
                col_name = df.columns[col_idx] if col_idx < len(df.columns) else f"Column_{col_idx}"
                row_data[col_name] = value
        
        # Add this row to our results if it has any non-None values
        if any(v is not None for v in row_data.values()):
            hierarchical_data.append(row_data)
        
        # Mark rows as processed
        for i in range(skip_rows):
            rows_processed.add(row_idx + i)
    
    # Create the final result structure with metadata and data
    result = {
        "metadata": metadata,
        "data": hierarchical_data
    }
    
    # Write to JSON
    with open(json_file, 'w') as f:
        json.dump(result, f, indent=2)
    
    print(f"Hierarchical data written to {json_file}")
    print(f"Processed {len(hierarchical_data)} hierarchical records with metadata")
    
    return result


def batch_process_excel_files(input_dir, output_dir):
    """
    Process all Excel files in a directory and convert them to JSON with metadata detection
    """
    input_path = Path(input_dir)
    output_path = Path(output_dir)
    os.makedirs(output_path, exist_ok=True)
    
    excel_files = list(input_path.glob("*.xlsx")) + list(input_path.glob("*.xls"))
    print(f"Found {len(excel_files)} Excel files to process")
    
    results = {}
    for excel_file in excel_files:
        json_file = output_path / f"{excel_file.stem}.json"
        try:
            result = convert_hierarchical_excel(str(excel_file), str(json_file))
            results[excel_file.name] = {
                "status": "success",
                "output_file": str(json_file),
                "metadata_rows": len(result["metadata"]),
                "data_rows": len(result["data"])
            }
        except Exception as e:
            print(f"Error processing {excel_file}: {str(e)}")
            results[excel_file.name] = {
                "status": "error",
                "error": str(e)
            }
    
    # Write summary report
    summary_file = output_path / "processing_summary.json"
    with open(summary_file, 'w') as f:
        json.dump(results, f, indent=2)
    
    return results


if __name__ == "__main__":
    # Example usage
    # result = convert_hierarchical_excel('hierarchical_data.xlsx', 'hierarchical_output.json')
    
    # For batch processing
    # results = batch_process_excel_files('data/structured', 'data/processed')
    
    print("Excel metadata processor module ready.") 