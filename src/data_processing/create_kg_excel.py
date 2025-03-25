from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

def create_knowledge_graph_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "KG_Research_Data"

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40

    # Add header
    ws.merge_cells('A1:E1')
    ws['A1'] = "Knowledge Graph: Research Collaboration Network"
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Add column headers
    headers = ["Entity ID", "Entity Type", "Entity Name", "Properties", "Relationships"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=header).font = Font(bold=True)

    # Data for entities
    entities = [
        {
            "id": "KG_001",
            "type": "Researcher",
            "name": "Dr. Jane Smith",
            "properties": [
                "PhD in Computer Science",
                "Specialization: AI",
                "Years of experience: 15"
            ],
            "relationships": [
                "Collaborates with KG_002",
                "Supervises KG_005, KG_006",
                "Member of Lab KG_010",
                "Published with KG_003"
            ]
        },
        {
            "id": "KG_002",
            "type": "Institution",
            "name": "Tech University",
            "properties": [
                "Founded: 1950",
                "Location: New York",
                "Research focus: AI, Robotics"
            ],
            "relationships": [
                "Employs KG_001",
                "Partners with KG_004",
                "Hosts Lab KG_010",
                "Funds project KG_015"
            ]
        }
    ]

    # Add entity data
    row = 3
    for entity in entities:
        ws.cell(row=row, column=1, value=entity['id'])
        ws.cell(row=row, column=2, value=entity['type'])
        ws.cell(row=row, column=3, value=entity['name'])

        # Properties
        prop_cell = ws.cell(row=row, column=4)
        prop_cell.value = "\n".join(f"- {prop}" for prop in entity['properties'])
        prop_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=row, start_column=4, end_row=row+len(entity['properties'])-1, end_column=4)

        # Relationships
        rel_cell = ws.cell(row=row, column=5)
        rel_cell.value = "\n".join(f"- {rel}" for rel in entity['relationships'])
        rel_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=row, start_column=5, end_row=row+len(entity['relationships'])-1, end_column=5)

        row += max(len(entity['properties']), len(entity['relationships']))

    # Save the workbook
    wb.save("KnowledgeGraph_Research_Data.xlsx")

if __name__ == "__main__":
    create_knowledge_graph_excel()
    print("Excel file 'KnowledgeGraph_Research_Data.xlsx' has been created.")
