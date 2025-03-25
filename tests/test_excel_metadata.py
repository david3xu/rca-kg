#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Test cases for Excel metadata detection.
"""

import unittest
import sys
import os
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, PatternFill

# Add parent directory to path to import modules
sys.path.append(str(Path(__file__).parent.parent))

from src.data_processing.excel_metadata_processor import convert_hierarchical_excel


class TestExcelMetadataProcessor(unittest.TestCase):
    """Test cases for the Excel metadata processor."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.test_data_dir = Path(__file__).parent / "test_data"
        os.makedirs(self.test_data_dir, exist_ok=True)
        
        # Create a test Excel file with metadata and hierarchical data
        self.test_excel_file = self.test_data_dir / "test_metadata.xlsx"
        self.test_json_file = self.test_data_dir / "test_metadata.json"
        
        # Create Excel file with metadata section
        self._create_test_excel_file()
    
    def _create_test_excel_file(self):
        """Create a test Excel file with metadata and hierarchical data."""
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # Add metadata section (typically at the top, with merged cells)
        sheet.merge_cells('A1:E1')
        sheet['A1'] = "TEST REPORT - EQUIPMENT FAILURE ANALYSIS"
        sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet['A1'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Add metadata rows
        sheet['A2'] = "Report Date:"
        sheet['B2'] = "2023-10-15"
        sheet['D2'] = "Report ID:"
        sheet['E2'] = "RCA-2023-001"
        
        sheet['A3'] = "Facility:"
        sheet['B3'] = "Plant A"
        sheet['D3'] = "Unit:"
        sheet['E3'] = "Processing Unit 3"
        
        # Add an empty row
        
        # Add header row (where main data starts)
        headers = ["Equipment", "Failure Mode", "Contributing Factor", "Impact", "Action Items"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=5, column=col).value = header
            sheet.cell(row=5, column=col).fill = PatternFill(
                start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        
        # Add hierarchical data with merged cells
        # Example: Equipment merged across multiple rows for different failure modes
        sheet.merge_cells('A6:A8')
        sheet['A6'] = "Pump XYZ-101"
        
        # First failure mode and details
        sheet['B6'] = "Bearing Failure"
        sheet['C6'] = "Inadequate Lubrication"
        sheet['D6'] = "Production Loss: 24 hours"
        sheet['E6'] = "Implement lubricant monitoring"
        
        # Second failure mode for same equipment
        sheet['B7'] = "Bearing Failure"
        sheet['C7'] = "Contamination"
        sheet['D7'] = "Production Loss: 12 hours"
        sheet['E7'] = "Improve sealing"
        
        # Third failure mode for same equipment
        sheet['B8'] = "Seal Leakage"
        sheet['C8'] = "Wear and Tear"
        sheet['D8'] = "Environmental Release"
        sheet['E8'] = "Replace with upgraded seal type"
        
        # Second equipment with its own failure modes
        sheet.merge_cells('A9:A10')
        sheet['A9'] = "Compressor ABC-102"
        
        # Failure modes for second equipment
        sheet['B9'] = "Valve Damage"
        sheet['C9'] = "Foreign Object"
        sheet['D9'] = "Production Loss: 48 hours"
        sheet['E9'] = "Install inlet filter"
        
        sheet['B10'] = "Vibration"
        sheet['C10'] = "Misalignment"
        sheet['D10'] = "Reduced Equipment Life"
        sheet['E10'] = "Implement laser alignment"
        
        # Save the file
        wb.save(self.test_excel_file)
        print(f"Created test Excel file at {self.test_excel_file}")
    
    def test_metadata_detection(self):
        """Test metadata detection in Excel file."""
        # Process the test Excel file
        result = convert_hierarchical_excel(str(self.test_excel_file), str(self.test_json_file))
        
        # Verify metadata section was detected
        self.assertIn("metadata", result)
        metadata = result["metadata"]
        
        # Check for the title in metadata
        title_found = False
        for key, value in metadata.items():
            if isinstance(value, str) and "TEST REPORT" in value:
                title_found = True
                break
            elif isinstance(value, dict):
                for inner_key, inner_value in value.items():
                    if isinstance(inner_value, str) and "TEST REPORT" in inner_value:
                        title_found = True
                        break
        
        self.assertTrue(title_found, "Title not found in metadata")
        
        # Verify data section contains hierarchical data
        self.assertIn("data", result)
        data = result["data"]
        
        # Should have at least 5 data rows
        self.assertGreaterEqual(len(data), 5, "Not enough data rows detected")
        
        # Check for hierarchical data structure (equipment with multiple failures)
        equipment_entries = set()
        for entry in data:
            for key, value in entry.items():
                if key.lower() in ("equipment", "column_0") and value is not None:
                    equipment_entries.add(value)
        
        # Should have found at least 2 different equipment entries
        self.assertGreaterEqual(len(equipment_entries), 2, 
                                "Not enough unique equipment entries found")
    
    def tearDown(self):
        """Clean up test fixtures."""
        # Normally we would clean up, but for debugging leave the files
        pass


if __name__ == "__main__":
    unittest.main() 